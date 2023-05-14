import numpy as np
import scipy.io as scio
import pickle
import math
import statistics
import os
import datetime
import scipy.signal as sciSignal
from utils import get_data, convert_time
from openpyxl import load_workbook
from saglrstepvar import saglrstepvar

fs = 1000

categories = ['单词', '句子', '静态表情', '动态表情', '无关动作']
# # words =['ahead','back','left','right','down','up','turn','on','off','over','out','happy','more','tired','afraid','hurt','pain','sad','fever','dizzy','numb','itchy','thirsty','full','hungry','cold','hot'
# #   ,'thank','sorry','sleepy','dyspnea','stop','pause','previous','next','favourite','ligten','dim','go' ,'get','good','better','bad','yes','no','ok','want','help','message','bath','poop','pee','time','wait','come','slow','emergency','pardon','wash','quiet','dressed','move','lie'
# #    ,'I','You','he','it','here','see','when','who','what','why','where','less','much','hi','body']
# words =['ahead','back','left','right','down','up','turn']
# # sentences =['Much better ','thank you','see you','i see ','Come on', 'Slow down .','Move body ','Turn over','Turn on ','Turn off','turn up','turn down','go off','go on','go out','wash up','get dressed','get off','lie down','get up']
# sentences =['Much better ','thank you','see you','i see ','Come on']
# tran_expressions = ['委屈（噘嘴）', '打哈欠', '难过（皱眉）', '咬牙', '咳嗽', '开心（微笑唇）', '静息']
# dynm_expressions =['微笑唇过渡到噘嘴','左鼓腮过渡到右鼓腮']
# irrelevant =['不发声，无关动作2min']
words =['ahead']
sentences =['Much better ','thank you','see you','i see ','Come on']#'Slow down .','Move body ','Turn over','Turn on ','Turn off','turn up','turn down','go off','go on','go out','wash up','get dressed','get off','lie down','get up']
tran_expressions = ['委屈（噘嘴）']
dynm_expressions =['微笑唇过渡到噘嘴'] 
irrelevant =['不发声，无关动作2min']
n_tasks = [len(words),len(sentences),len(tran_expressions),len(dynm_expressions),len(irrelevant)]
task_duration = [2,3,2,3,10]
n_channels = 64
n_trials = 3
n_categories = 5
n_total = sum(n_tasks[:-1])*n_trials+1
task_cumsum = np.cumsum(np.array(n_tasks[:-1] * n_trials+[1]))
 
task_duration_vec = [[task_duration[i]]*n_tasks[i] for i in range(len(categories)-1)]*n_trials + [[task_duration[-1]]]
task_duration_vec = [item for sublist in task_duration_vec for item in sublist]

rest_duration_vec = []
for i in range(len(categories)-1):
    for j in range(n_tasks[i]):
        if j % n_tasks[i] == 0: # 组间
            rest_duration_vec.append(3)
        else:
            rest_duration_vec.append(1) # 组内
rest_duration_vec = rest_duration_vec * n_trials
rest_duration_vec.append(3) # 无关动作


reaction_time = 0.25*1000

path = './DataSave/'
subject = ''
date = str(datetime.date.today())
mode = 'SS'



def label_process(path, subject, date, mode):
    # 实验开始的时间戳
    EventStart_timestamp = np.zeros(n_total)
    # 接收到数据包的时间戳
    PackTrig_timestamp = []
    # 记录各trial开始时间的文件
    EventTrigname = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_PRlog.xlsx'
    # 记录收到各包时间的文件
    PackTrigname = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_timestampData.txt'
    # 记录随机label的文件
    RandLabelname = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_PRlog.mat'
    # 人工添加的trigger
    trigger = []
    # 取出trial开始的时间
    wb = load_workbook(EventTrigname)
    ws = wb[wb.sheetnames[0]]
    rows = ws.rows
    content = []
    for row in rows:
        content.append([col.value for col in row])
    for i in range(n_total):
        each = str(content[1+2*i][1])
        temp = int(each[0:2]) * 3600 * 1000 + int(each[2:4]) * 60 * 1000 + int(each[4:6]) * 1000 + int(each[6:])
        EventStart_timestamp[i] = temp
    # 遍历每个包的到达时间，并记录到temp_packtrig中
    with open(PackTrigname) as f:
        file = f.readlines()
    for each in file:
        each = each.strip('\n').split(' ')
        if 'x' not in each[0]:
            temp = convert_time(each[0])
        PackTrig_timestamp.append(temp)
        trigger.append(int(each[2]))

    print('--------------')
    print('Expected amount of packs')
    print((PackTrig_timestamp[-1]-PackTrig_timestamp[0])//20)
    print('--------------')

    # 随机数order
    file = scio.loadmat(RandLabelname)
    randIndex = file['randIndex'][0]
    # 设置list记录某一trial(除无关）估计的起始时间，结束时间，以及randIndex
    TrialData = np.zeros([n_total,3])
    for i in range(n_total):
        duration = task_duration_vec[i] * 1000
        # 250 ms 反应时间
        data_be = EventStart_timestamp[i] - PackTrig_timestamp[0] + reaction_time
        data_be = int(data_be // 100 * 100)
        data_en = int(data_be + duration)
        # 保存起始时间，结束时间，randIndex（无关设为-1）
        TrialData[i][0] = data_be
        TrialData[i][1] = data_en
        if i < n_total - 1:
            TrialData[i][2] = randIndex[i]
        else:
            TrialData[i][2] = -1
        
    print('--------------')
    print('trial start and end time')
    print(TrialData[:,:-1])
    print('--------------')
    # 保存TrialData
    filename = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_TrialData.pkl'
    with open(filename, 'wb') as f:
        pickle.dump(TrialData, f)
    # 保存trigger
    filename = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_trigger.pkl'
    with open(filename, 'wb') as f:
        pickle.dump(trigger, f)
        

def data_process_auto(path, subject, date, mode):
    if not os.path.exists(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat'):
        # 原始byte数据
        unseg_dataname = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_EMGData.bin'
        with open(unseg_dataname, 'rb') as f:
            packs = f.read()
        # 采样点数目，每个采样点为64维度，4byte的浮点数
        n_sample = len(packs)//4//n_channels
        EMGData = np.zeros([n_sample, n_channels])
        for i in range(n_sample):
            for j in range(n_channels):
                EMGData[i,j] = get_data(packs[i*256+j*4:i*256+j*4+4])
        os.makedirs(path + subject + '_' + date + '_' + mode + '/Data/', exist_ok=True)
        scio.savemat(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat', {'data':EMGData})
    else:
        EMGData = scio.loadmat(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat')['data']
    print('--------------')
    print('emg data shape')
    print(EMGData.shape)
    print('Actual amount of packs (including dummy data): ')
    print(EMGData.shape[0]//20)
    print('--------------')

    # 对每个通道的信号进行10-499.5HZ的带通滤波以及50Hz的陷波滤波
    sos = sciSignal.butter(8, [10*2/fs, 499.5*2/fs], 'bandpass', output='sos')
    nb, na = sciSignal.iirnotch(50*2/fs, 30.0)
    nsos = sciSignal.tf2sos(nb, na)
    EMGData = sciSignal.sosfiltfilt(sos, EMGData, axis=0)
    EMGData = sciSignal.sosfiltfilt(nsos, EMGData, axis=0)

    # 读取每个trial的时间节点以及label
    TrialDataName = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_TrialData.pkl'
    with open(TrialDataName, 'rb') as f:
        TrialData = pickle.load(f)

    # 估计肌电信号的开始位置
    EstimatedBurstTime = np.zeros([TrialData.shape[0], n_channels])
    for i in range(n_channels):
        emgdata_ch = EMGData[:,i]
        [t0,_,_,ga] = saglrstepvar(emgdata_ch, 200, 100, 100 ,'multiple', 1)
        gt = ga[t0]
        dis = TrialData[:,0].reshape(-1,1) - t0.reshape(1,-1)
        loc = np.argmin(abs(dis), axis=0)
        uniqLoc = np.unique(loc)
        for j in range(len(uniqLoc)):
            temp1 = np.where(loc==uniqLoc[j])
            temp = temp1[0]
            locS = np.argsort(gt[temp])
            if len(locS) > 1:
                tempT0 = t0[temp[locS[0:2]]]
            else:
                tempT0 = t0[temp[locS]]
            finT0 = np.argmin(abs(tempT0 - np.tile(TrialData[j,0], (1, len(tempT0)))))
            EstimatedBurstTime[uniqLoc[j], i] = tempT0[finT0]
    
    idx1 = ((EstimatedBurstTime - TrialData[:,1].reshape(-1,1)) > 0)
    duration_vec = np.array([task_duration_vec]).reshape(-1,1) * 1000
    idx2 = (abs(EstimatedBurstTime - TrialData[:,0].reshape(-1,1)) > duration_vec)
    EstimatedBurstTime[idx1] = np.nan
    EstimatedBurstTime[idx2] = np.nan
    print(EstimatedBurstTime)


    # 100ms精度
    EstimatedBurstTime_tmp = np.floor(EstimatedBurstTime / 100)
    finBurstTime = np.zeros(n_total)
    WrongTrial = []
    for i in range(n_total):
        valid_start = [t for t in EstimatedBurstTime_tmp[i,:] if not math.isnan(t)]
        if len(valid_start):
            finBurstTime[i] = statistics.mode(valid_start)*100
        else:
            print('You may speak at wrong time(earlier or later) in trial ' + str(i+1))
            WrongTrial.append(i)
    
    print('--------------')
    print('actual data start and end time')
    # 分组记录数据
    finBurstTime = finBurstTime.astype(int)
    emgdata = []
    for i in range(n_total):
        # 在任务数据中添加一点休息时候的数据，让分类器更鲁棒
        spk_be = finBurstTime[i] - 100
        # 数据的开始和结束时间（包含rest）
        data_be = spk_be - rest_duration_vec[i] * 1000
        data_en = spk_be + task_duration_vec[i] * 1000
        print(spk_be,data_en)
        if data_be < 0 or data_en > EMGData.shape[0]:
            data_be = 0
            data_en = 0
        emgdata.append(EMGData[int(data_be):int(data_en),:])

    
    filename = path + subject + '_' + date + '_' + mode + '/Data/' + 'emg_whole.mat'
    scio.savemat(filename, {'EMGData':np.array(emgdata,dtype=object)})
    print('--------------')

    # 取出任务和休息的数据，并按任务类别存放
    emg_inter_group_rest = [[] for i in range(len(categories))] # 组间休息
    emg_intra_group_rest = [[] for i in range(len(categories)-1)] # 组内休息
    emg_active = [[] for i in range(len(categories))] # 活动数据
    labels = [[] for i in range(len(categories)-1)] # 数据标签
    for i in range(n_total):
        if i in WrongTrial:
            continue
        if i < n_total - 1:
            i_cat = np.argmin(i+1>task_cumsum) % 4
        else:
            i_cat = 4
        temp = emgdata[i]
        emg_active[i_cat].append(temp[-task_duration_vec[i]*1000:,:])
        if rest_duration_vec[i] == 3:
            emg_inter_group_rest[i_cat].append(temp[:rest_duration_vec[i]*1000,:])
        else:
            emg_intra_group_rest[i_cat].append(temp[:rest_duration_vec[i]*1000,:])
        if i_cat != 4:
            labels[i_cat].append(TrialData[i,-1])
    # 保存数据
    # 单词
    dir = path + subject + '_' + date + '_' + mode + '/Data/'+ '/word/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[0]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[0]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[0]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[0])})

    # 句子
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/sentence/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[1]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[1]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[1]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[1])})

    # 静态表情
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/trans_expression/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[2]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[2]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[2]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[2])})

    # 动态表情
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/dynm_expression/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[3]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[3]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[3]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[3])})

    # 无关动作
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/irrelevant/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[4]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[4]})

    # 有问题的trial
    filename = path + subject + '_' + date + '_' + mode + '/Data/' + 'wrong_trial.mat'
    scio.savemat(filename,{'wrong_trial':WrongTrial})

def data_process_manual(path, subject, date, mode):
    if not os.path.exists(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat'):
        # 原始byte数据
        unseg_dataname = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_EMGData.bin'
        with open(unseg_dataname, 'rb') as f:
            packs = f.read()
        # 采样点数目，每个采样点为64维度，4byte的浮点数
        n_sample = len(packs)//4//n_channels
        EMGData = np.zeros([n_sample, n_channels])
        for i in range(n_sample):
            for j in range(n_channels):
                EMGData[i,j] = get_data(packs[i*256+j*4:i*256+j*4+4])
        os.makedirs(path + subject + '_' + date + '_' + mode + '/Data/', exist_ok=True)
        scio.savemat(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat', {'data':EMGData})
    else:
        EMGData = scio.loadmat(path + subject + '_' + date + '_' + mode + '/Data/' + 'EMG_test.mat')['data']
    print('--------------')
    print('emg data shape')
    print(EMGData.shape)
    print('Actual amount of packs (including dummy data): ')
    print(EMGData.shape[0]//20)
    print('--------------')

    # 对每个通道的信号进行10-499.5HZ的带通滤波以及50Hz的陷波滤波
    sos = sciSignal.butter(8, [10*2/fs, 499.5*2/fs], 'bandpass', output='sos')
    nb, na = sciSignal.iirnotch(50*2/fs, 30.0)
    nsos = sciSignal.tf2sos(nb, na)
    EMGData = sciSignal.sosfiltfilt(sos, EMGData, axis=0)
    EMGData = sciSignal.sosfiltfilt(nsos, EMGData, axis=0)
    
    # 读取手动设置的trigger
    
    filename = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_trigger.pkl'
    with open(filename, 'rb') as f:
        trigger = pickle.load(f)

    # 读取label
    TrialDataName = path + subject + '_' + date + '_' + mode + '/' + subject + '_' + date + '_TrialData.pkl'
    with open(TrialDataName, 'rb') as f:
        TrialData = pickle.load(f)


    # 取出任务和休息的数据，并按任务类别存放
    emgdata = []
    emg_inter_group_rest = [[] for i in range(len(categories))] # 组间休息
    emg_intra_group_rest = [[] for i in range(len(categories)-1)] # 组内休息
    emg_active = [[] for i in range(len(categories))] # 活动数据
    labels = [[] for i in range(len(categories)-1)] # 数据标签

    # 根据trigger划分数据
    i = 0
    i_trigger = 0
    while i < n_total:
        if i < n_total - 1:
            i_cat = np.argmin(i+1>task_cumsum) % 4
        else:
            i_cat = 4
        n_packs = task_duration[i_cat] * 1000 // 20
        pack_cnt = 0
        if trigger[i_trigger] == 0 or trigger[i_trigger] == -1:
            i_trigger += 1
        else:
            st = i_trigger
            valid_flag = 0
            while trigger[i_trigger] == 1:
                pack_cnt += 1
                i_trigger += 1
                if pack_cnt == n_packs:
                    valid_flag = 1
                    break
            
            if valid_flag == 1:
                emgdata.append(EMGData[(st-1-rest_duration_vec[i]*1000//20)*20:(st-1+n_packs)*20,:])
                print((st-1)*20,(st-1+n_packs)*20)
                emg_active[i_cat].append(EMGData[(st-1)*20:(st-1+n_packs)*20,:])
                if rest_duration_vec[i] == 3:
                    emg_inter_group_rest[i_cat].append(EMGData[(st-1-150)*20:(st-1)*20,:])
                else:
                    emg_intra_group_rest[i_cat].append(EMGData[(st-1-50)*20:(st-1)*20,:])
                if i_cat != 4:
                    labels[i_cat].append(TrialData[i,-1])
                i += 1

    # 保存全部数据

    filename = path + subject + '_' + date + '_' + mode + '/Data/' + 'emg_whole.mat'
    scio.savemat(filename, {'EMGData':np.array(emgdata,dtype=object)})

    # 保存数据
    # 单词
    dir = path + subject + '_' + date + '_' + mode + '/Data/'+ '/word/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[0]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[0]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[0]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[0])})

    # 句子
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/sentence/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[1]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[1]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[1]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[1])})

    # 静态表情
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/trans_expression/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[2]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[2]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[2]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[2])})

    # 动态表情
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/dynm_expression/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[3]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[3]})
    filename = dir + 'emg_intragroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_intra_group_rest[3]})
    filename = dir + 'label.mat'
    scio.savemat(filename, {'label':np.array(labels[3])})

    # 无关动作
    dir = path + subject + '_' + date + '_' + mode + '/Data/' + '/irrelevant/'
    os.makedirs(dir, exist_ok=True)
    filename = dir + 'emg_active.mat'
    scio.savemat(filename, {'EMGData':emg_active[4]})
    filename = dir + 'emg_intergroup_rest.mat'
    scio.savemat(filename, {'EMGData':emg_inter_group_rest[4]})



label_process(path, subject, date, mode)
# data_process_auto(path, subject, date, mode)
data_process_manual(path, subject, date, mode)
